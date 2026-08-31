import io
import os
import re
import json
import time
import copy
from pathlib import Path

import streamlit as st
from PIL import Image, ImageDraw, ImageFont

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries

from google import genai
from google.genai import types


# ============================================================
# STREAMLIT CONFIG
# ============================================================

st.set_page_config(
    page_title="Gemini AI - Dịch Trung Việt",
    page_icon="🌐",
    layout="wide",
)


# ============================================================
# CONSTANTS
# ============================================================

PRIMARY_MODEL = "gemini-2.5-flash"
FALLBACK_MODEL = "gemini-2.5-flash-lite"

SUPPORTED_EXTENSIONS = [
    "xlsx",
    "xlsm",
    "png",
    "jpg",
    "jpeg",
    "webp",
    "bmp",
]

IMAGE_EXTENSIONS = {
    "png",
    "jpg",
    "jpeg",
    "webp",
    "bmp",
}

EXCEL_EXTENSIONS = {
    "xlsx",
    "xlsm",
}

MAX_TEXT_BATCH = 40
MAX_RETRIES = 3


# ============================================================
# CSS
# ============================================================

st.markdown(
    """
    <style>
    .main-title {
        font-size: 34px;
        font-weight: 800;
        margin-bottom: 4px;
    }

    .sub-title {
        color: #666;
        font-size: 15px;
        margin-bottom: 20px;
    }

    .rule-box {
        border: 1px solid #d8e2f0;
        background: #f6f9ff;
        padding: 15px;
        border-radius: 10px;
        margin: 10px 0 20px 0;
    }

    .success-box {
        border: 1px solid #cce8d2;
        background: #f3fff5;
        padding: 15px;
        border-radius: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# HEADER
# ============================================================

st.markdown(
    '<div class="main-title">🌐 Gemini AI - Dịch Trung ↔ Việt</div>',
    unsafe_allow_html=True,
)

st.markdown(
    '<div class="sub-title">'
    "Dịch Excel và hình ảnh bằng Gemini AI"
    "</div>",
    unsafe_allow_html=True,
)


# ============================================================
# SIDEBAR
# ============================================================

st.sidebar.header("⚙️ Cấu hình")

direction = st.sidebar.selectbox(
    "Chọn hướng dịch",
    [
        "Trung → Việt",
        "Việt → Trung",
    ],
)

st.sidebar.markdown("---")

st.sidebar.subheader("🤖 Gemini AI")

api_key_input = st.sidebar.text_input(
    "Gemini API Key",
    type="password",
    placeholder="AIza...",
    help=(
        "Có thể nhập trực tiếp hoặc để trống nếu đã cấu hình "
        "GEMINI_API_KEY trong Streamlit Secrets."
    ),
)

model_choice = st.sidebar.selectbox(
    "Model chính",
    [
        "gemini-2.5-flash",
        "gemini-2.5-flash-lite",
    ],
    index=0,
)

if direction == "Trung → Việt":
    st.sidebar.success(
        "Kết quả:\n\n"
        "🇨🇳 Tiếng Trung\n"
        "🇻🇳 Tiếng Việt"
    )
else:
    st.sidebar.success(
        "Kết quả:\n\n"
        "🇨🇳 Tiếng Trung\n"
        "🇻🇳 Tiếng Việt"
    )

st.sidebar.markdown(
    """
    <div class="rule-box">
    <b>Quy tắc bắt buộc:</b><br><br>
    🇨🇳 Tiếng Trung luôn ở trên<br>
    🇻🇳 Tiếng Việt luôn ở ngay bên dưới
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# API KEY
# ============================================================

def get_api_key():
    # 1. Sidebar
    if api_key_input and api_key_input.strip():
        return api_key_input.strip()

    # 2. Streamlit Secrets
    try:
        key = st.secrets.get("GEMINI_API_KEY")

        if key:
            return str(key).strip()

    except Exception:
        pass

    # 3. Environment
    key = os.getenv("GEMINI_API_KEY")

    if key:
        return key.strip()

    return ""


API_KEY = get_api_key()


# ============================================================
# GEMINI CLIENT
# ============================================================

@st.cache_resource(show_spinner=False)
def create_gemini_client(api_key):
    if not api_key:
        return None

    return genai.Client(api_key=api_key)


client = create_gemini_client(API_KEY)


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_text(value):
    if value is None:
        return ""

    return str(value).strip()


def is_formula(value):
    return (
        isinstance(value, str)
        and value.startswith("=")
    )


def is_number(value):
    if value is None:
        return False

    if isinstance(value, (int, float)):
        return True

    text = str(value).strip()

    if not text:
        return False

    patterns = [
        r"^-?\d+$",
        r"^-?\d+\.\d+$",
        r"^-?\d+,\d+$",
        r"^-?\d{1,3}(,\d{3})+(\.\d+)?$",
        r"^-?\d{1,3}(\.\d{3})+(,\d+)?$",
        r"^-?\d+(\.\d+)?%$",
    ]

    return any(
        re.fullmatch(pattern, text)
        for pattern in patterns
    )


def looks_like_code_or_id(text):
    if not text:
        return False

    text = str(text).strip()

    if len(text) <= 2:
        return False

    # Ví dụ:
    # ABC-123
    # M12-001
    # PART_12345
    if re.fullmatch(
        r"[A-Z]{2,10}[-_/]?\d{2,}",
        text,
    ):
        return True

    # Mã kỹ thuật toàn chữ/số
    if re.fullmatch(
        r"[A-Z0-9_-]{5,}",
        text,
    ):
        return True

    return False


def should_translate_cell(value):
    """
    Chỉ dịch nội dung văn bản thực sự.

    Không dịch:
    - ô rỗng
    - số
    - boolean
    - công thức
    - mã kỹ thuật rõ ràng
    """

    if value is None:
        return False

    if isinstance(value, bool):
        return False

    if isinstance(value, (int, float)):
        return False

    text = str(value).strip()

    if not text:
        return False

    if is_formula(text):
        return False

    if is_number(text):
        return False

    if looks_like_code_or_id(text):
        return False

    return True


# ============================================================
# JSON PARSER
# ============================================================

def extract_json(text):
    if not text:
        raise ValueError(
            "Gemini không trả về dữ liệu."
        )

    text = text.strip()

    # JSON trực tiếp
    try:
        return json.loads(text)
    except Exception:
        pass

    # Markdown JSON
    text = re.sub(
        r"^```json\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )

    text = re.sub(
        r"\s*```$",
        "",
        text,
    )

    try:
        return json.loads(text)
    except Exception:
        pass

    # Tìm object
    first_obj = text.find("{")
    last_obj = text.rfind("}")

    if (
        first_obj >= 0
        and last_obj > first_obj
    ):
        candidate = text[
            first_obj:last_obj + 1
        ]

        try:
            return json.loads(candidate)
        except Exception:
            pass

    # Tìm array
    first_list = text.find("[")
    last_list = text.rfind("]")

    if (
        first_list >= 0
        and last_list > first_list
    ):
        candidate = text[
            first_list:last_list + 1
        ]

        try:
            return json.loads(candidate)
        except Exception:
            pass

    raise ValueError(
        "Không thể đọc JSON từ phản hồi Gemini."
    )


# ============================================================
# GEMINI CALL
# ============================================================

def gemini_generate(
    prompt,
    model,
    image=None,
    mime_type=None,
):
    """
    Gọi Gemini có retry.

    Nếu API lỗi tạm thời:
    503 / 429 / timeout / overloaded...
    sẽ tự thử lại.
    """

    if client is None:
        raise RuntimeError(
            "Chưa cấu hình GEMINI_API_KEY."
        )

    last_error = None

    for attempt in range(
        1,
        MAX_RETRIES + 1,
    ):
        try:

            if image is None:

                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        temperature=0,
                        response_mime_type="application/json",
                    ),
                )

            else:

                image_part = types.Part.from_bytes(
                    data=image,
                    mime_type=mime_type,
                )

                response = client.models.generate_content(
                    model=model,
                    contents=[
                        prompt,
                        image_part,
                    ],
                    config=types.GenerateContentConfig(
                        temperature=0,
                        response_mime_type="application/json",
                    ),
                )

            if (
                response is None
                or not response.text
            ):
                raise RuntimeError(
                    "Gemini trả về phản hồi rỗng."
                )

            return response.text, model

        except Exception as exc:

            last_error = exc

            error_text = str(exc).lower()

            temporary_error = any(
                keyword in error_text
                for keyword in [
                    "503",
                    "unavailable",
                    "429",
                    "resource exhausted",
                    "timeout",
                    "deadline",
                    "internal",
                    "overloaded",
                ]
            )

            if (
                attempt < MAX_RETRIES
                and temporary_error
            ):
                time.sleep(2 * attempt)
                continue

            raise last_error


# ============================================================
# TRANSLATE BATCH
# ============================================================

def translate_batch_with_gemini(
    items,
):
    """
    items:

    [
        {
            "id": 1,
            "text": "产品名称"
        }
    ]

    trả:

    {
        1: "Tên sản phẩm"
    }
    """

    if not items:
        return {}, PRIMARY_MODEL

    if direction == "Trung → Việt":

        source_name = "Chinese"
        target_name = "Vietnamese"

    else:

        source_name = "Vietnamese"
        target_name = "Chinese"

    payload = json.dumps(
        items,
        ensure_ascii=False,
    )

    prompt = f"""
You are a professional Chinese-Vietnamese translator.

Translation direction:
{source_name} -> {target_name}

Translate ONLY the text values.

IMPORTANT RULES:

1. Preserve every ID exactly.
2. Every input ID must appear exactly once.
3. Do not omit any item.
4. Do not merge items.
5. Do not add explanations.
6. Do not summarize.
7. Preserve numbers exactly.
8. Preserve units exactly.
9. Preserve model numbers exactly.
10. Preserve technical terminology.
11. Preserve manufacturing terminology.
12. Preserve punctuation where appropriate.
13. Preserve line breaks when possible.
14. If the text is already in the target language, return it unchanged.
15. Return JSON only.

Input:
{payload}

Required output format:

{{
  "translations": [
    {{
      "id": 1,
      "translation": "..."
    }}
  ]
}}
"""

    primary_error = None

    # ========================================================
    # PRIMARY
    # ========================================================

    try:

        raw, used_model = gemini_generate(
            prompt=prompt,
            model=model_choice,
        )

        data = extract_json(raw)

        translations = data.get(
            "translations",
            [],
        )

        result = {}

        for item in translations:

            if not isinstance(
                item,
                dict,
            ):
                continue

            item_id = item.get("id")
            translated = item.get(
                "translation",
                "",
            )

            if item_id is None:
                continue

            try:
                item_id = int(item_id)
            except Exception:
                continue

            result[item_id] = (
                str(translated)
                if translated is not None
                else ""
            )

        expected_ids = {
            int(item["id"])
            for item in items
        }

        returned_ids = set(result.keys())

        if expected_ids != returned_ids:

            missing = (
                expected_ids
                - returned_ids
            )

            raise ValueError(
                f"Gemini trả thiếu ID: {missing}"
            )

        return result, used_model

    except Exception as exc:

        primary_error = exc


    # ========================================================
    # FALLBACK
    # ========================================================

    fallback_model = (
        FALLBACK_MODEL
        if model_choice
        != FALLBACK_MODEL
        else PRIMARY_MODEL
    )

    try:

        raw, used_model = gemini_generate(
            prompt=prompt,
            model=fallback_model,
        )

        data = extract_json(raw)

        translations = data.get(
            "translations",
            [],
        )

        result = {}

        for item in translations:

            if not isinstance(
                item,
                dict,
            ):
                continue

            item_id = item.get("id")
            translated = item.get(
                "translation",
                "",
            )

            if item_id is None:
                continue

            try:
                item_id = int(item_id)
            except Exception:
                continue

            result[item_id] = (
                str(translated)
                if translated is not None
                else ""
            )

        expected_ids = {
            int(item["id"])
            for item in items
        }

        returned_ids = set(result.keys())

        if expected_ids != returned_ids:

            missing = (
                expected_ids
                - returned_ids
            )

            raise ValueError(
                f"Fallback trả thiếu ID: {missing}"
            )

        return result, used_model

    except Exception as fallback_error:

        raise RuntimeError(
            "Gemini chính và Gemini fallback đều thất bại.\n\n"
            f"Primary: {primary_error}\n\n"
            f"Fallback: {fallback_error}"
        )


# ============================================================
# EXCEL - STYLE
# ============================================================

def copy_cell_style(src, dst):
    """
    Copy toàn bộ style cơ bản.
    """

    if isinstance(
        src,
        MergedCell,
    ):
        return

    if isinstance(
        dst,
        MergedCell,
    ):
        return

    if src.has_style:
        dst._style = copy.copy(
            src._style
        )

    if src.number_format:
        dst.number_format = (
            src.number_format
        )

    if src.protection:
        dst.protection = copy.copy(
            src.protection
        )

    if src.alignment:
        dst.alignment = copy.copy(
            src.alignment
        )

    if src.font:
        dst.font = copy.copy(
            src.font
        )

    if src.fill:
        dst.fill = copy.copy(
            src.fill
        )

    if src.border:
        dst.border = copy.copy(
            src.border
        )


def copy_row_dimension(
    ws,
    source_row,
    target_row,
):
    src = ws.row_dimensions[
        source_row
    ]

    dst = ws.row_dimensions[
        target_row
    ]

    if src.height is not None:
        dst.height = src.height

    dst.hidden = src.hidden
    dst.outlineLevel = (
        src.outlineLevel
    )
    dst.collapsed = src.collapsed


def copy_column_dimensions(
    ws,
):
    result = {}

    for key, dim in (
        ws.column_dimensions.items()
    ):

        result[key] = {
            "width": dim.width,
            "hidden": dim.hidden,
            "bestFit": dim.bestFit,
            "outlineLevel": (
                dim.outlineLevel
            ),
            "collapsed": dim.collapsed,
        }

    return result


def restore_column_dimensions(
    ws,
    data,
):
    for key, values in data.items():

        dim = ws.column_dimensions[
            key
        ]

        dim.width = values[
            "width"
        ]

        dim.hidden = values[
            "hidden"
        ]

        dim.bestFit = values[
            "bestFit"
        ]

        dim.outlineLevel = values[
            "outlineLevel"
        ]

        dim.collapsed = values[
            "collapsed"
        ]


# ============================================================
# EXCEL - COLLECT TEXT
# ============================================================

def collect_excel_translation_items(
    wb,
):
    items = []
    counter = 1

    for ws in wb.worksheets:

        for row in ws.iter_rows():

            for cell in row:

                if isinstance(
                    cell,
                    MergedCell,
                ):
                    continue

                value = cell.value

                if not should_translate_cell(
                    value
                ):
                    continue

                items.append(
                    {
                        "id": counter,
                        "sheet": ws.title,
                        "row": cell.row,
                        "column": cell.column,
                        "text": str(value),
                    }
                )

                counter += 1

    return items


# ============================================================
# EXCEL - TRANSLATE
# ============================================================

def translate_excel_texts(
    wb,
    progress_callback=None,
):
    all_items = (
        collect_excel_translation_items(
            wb
        )
    )

    if not all_items:
        return (
            {},
            PRIMARY_MODEL,
            0,
        )

    translations = {}

    total = len(all_items)

    batches = [
        all_items[
            i:i + MAX_TEXT_BATCH
        ]
        for i in range(
            0,
            total,
            MAX_TEXT_BATCH,
        )
    ]

    used_models = []

    for index, batch in enumerate(
        batches
    ):

        batch_input = [
            {
                "id": item["id"],
                "text": item["text"],
            }
            for item in batch
        ]

        batch_result, used_model = (
            translate_batch_with_gemini(
                batch_input
            )
        )

        translations.update(
            batch_result
        )

        used_models.append(
            used_model
        )

        if progress_callback:

            progress_callback(
                min(
                    1.0,
                    (index + 1)
                    / len(batches),
                )
            )

    if used_models:

        used_model_final = max(
            set(used_models),
            key=used_models.count,
        )

    else:

        used_model_final = (
            PRIMARY_MODEL
        )

    return (
        translations,
        used_model_final,
        total,
    )


# ============================================================
# EXCEL - BUILD TRANSLATION ROWS
# ============================================================

def apply_excel_translations(
    wb,
    translation_map,
    original_items,
):
    """
    Quy tắc:

    Trung -> Việt

        dòng gốc:
        Trung

        dòng mới:
        Việt


    Việt -> Trung

        dòng mới:
        Trung

        dòng gốc:
        Việt


    Vì vậy kết quả cuối luôn:

        Trung
        Việt
    """

    rows_by_sheet = {}

    for item in original_items:

        item_id = item["id"]

        if item_id not in translation_map:
            continue

        sheet = item["sheet"]
        row = item["row"]

        rows_by_sheet.setdefault(
            sheet,
            set(),
        ).add(row)

    # --------------------------------------------------------
    # Xử lý từng sheet
    # --------------------------------------------------------

    for ws in wb.worksheets:

        target_rows = rows_by_sheet.get(
            ws.title,
            set(),
        )

        if not target_rows:
            continue

        original_max_col = ws.max_column

        column_dimensions = (
            copy_column_dimensions(ws)
        )

        # Từ dưới lên
        for row in sorted(
            target_rows,
            reverse=True,
        ):

            # ------------------------------------------------
            # Lưu merge liên quan
            # ------------------------------------------------

            merge_ranges = []

            for merged in list(
                ws.merged_cells.ranges
            ):

                min_col, min_row, max_col, max_row = (
                    range_boundaries(
                        str(merged)
                    )
                )

                if (
                    min_row <= row
                    <= max_row
                ):

                    merge_ranges.append(
                        (
                            min_col,
                            min_row,
                            max_col,
                            max_row,
                        )
                    )

            # ------------------------------------------------
            # Unmerge các merge ảnh hưởng
            # ------------------------------------------------

            for merged in list(
                ws.merged_cells.ranges
            ):

                min_col, min_row, max_col, max_row = (
                    range_boundaries(
                        str(merged)
                    )
                )

                if (
                    min_row <= row
                    <= max_row
                ):

                    try:
                        ws.unmerge_cells(
                            str(merged)
                        )
                    except Exception:
                        pass

            # ------------------------------------------------
            # Lưu dữ liệu + style dòng gốc
            # ------------------------------------------------

            source_values = {}
            source_styles = {}
            source_hyperlinks = {}
            source_comments = {}

            for col in range(
                1,
                original_max_col + 1,
            ):

                cell = ws.cell(
                    row,
                    col,
                )

                source_values[col] = (
                    cell.value
                )

                source_styles[col] = (
                    copy.copy(
                        cell._style
                    )
                    if cell.has_style
                    else None
                )

                source_hyperlinks[col] = (
                    copy.copy(
                        cell.hyperlink
                    )
                    if cell.hyperlink
                    else None
                )

                source_comments[col] = (
                    copy.copy(
                        cell.comment
                    )
                    if cell.comment
                    else None
                )

            source_height = (
                ws.row_dimensions[
                    row
                ].height
            )

            source_hidden = (
                ws.row_dimensions[
                    row
                ].hidden
            )

            source_outline = (
                ws.row_dimensions[
                    row
                ].outlineLevel
            )

            source_collapsed = (
                ws.row_dimensions[
                    row
                ].collapsed
            )

            # ------------------------------------------------
            # Chèn dòng
            # ------------------------------------------------

            ws.insert_rows(
                row + 1,
                amount=1,
            )

            translated_row = row + 1

            # ------------------------------------------------
            # Copy row dimension
            # ------------------------------------------------

            if source_height is not None:

                ws.row_dimensions[
                    translated_row
                ].height = source_height

            ws.row_dimensions[
                translated_row
            ].hidden = source_hidden

            ws.row_dimensions[
                translated_row
            ].outlineLevel = (
                source_outline
            )

            ws.row_dimensions[
                translated_row
            ].collapsed = (
                source_collapsed
            )

            # ------------------------------------------------
            # Copy style/content ban đầu
            # ------------------------------------------------

            for col in range(
                1,
                original_max_col + 1,
            ):

                dst = ws.cell(
                    translated_row,
                    col,
                )

                if (
                    source_styles[col]
                    is not None
                ):

                    dst._style = copy.copy(
                        source_styles[col]
                    )

                dst.value = (
                    source_values[col]
                )

                if (
                    source_hyperlinks[col]
                    is not None
                ):

                    dst._hyperlink = (
                        copy.copy(
                            source_hyperlinks[
                                col
                            ]
                        )
                    )

                if (
                    source_comments[col]
                    is not None
                ):

                    dst.comment = (
                        copy.copy(
                            source_comments[
                                col
                            ]
                        )
                    )

            # ------------------------------------------------
            # Lấy các item thuộc dòng này
            # ------------------------------------------------

            row_items = [
                item
                for item in original_items
                if (
                    item["sheet"]
                    == ws.title
                    and item["row"]
                    == row
                )
            ]

            # ------------------------------------------------
            # Ghi bản dịch
            # ------------------------------------------------

            for item in row_items:

                item_id = item["id"]

                if (
                    item_id
                    not in translation_map
                ):
                    continue

                col = item["column"]

                original_text = (
                    item["text"]
                )

                translated_text = (
                    translation_map[
                        item_id
                    ]
                )

                source_cell = ws.cell(
                    row,
                    col,
                )

                target_cell = ws.cell(
                    translated_row,
                    col,
                )

                if direction == "Trung → Việt":

                    # Trên = Trung gốc
                    source_cell.value = (
                        original_text
                    )

                    # Dưới = Việt
                    target_cell.value = (
                        translated_text
                    )

                else:

                    # Trên = Trung dịch
                    source_cell.value = (
                        translated_text
                    )

                    # Dưới = Việt gốc
                    target_cell.value = (
                        original_text
                    )

            # ------------------------------------------------
            # Khôi phục merge ngang
            # ------------------------------------------------

            for (
                min_col,
                min_row,
                max_col,
                max_row,
            ) in merge_ranges:

                # Merge nằm đúng 1 dòng
                if (
                    min_row == row
                    and max_row == row
                ):

                    original_range = (
                        f"{get_column_letter(min_col)}"
                        f"{row}:"
                        f"{get_column_letter(max_col)}"
                        f"{row}"
                    )

                    translation_range = (
                        f"{get_column_letter(min_col)}"
                        f"{translated_row}:"
                        f"{get_column_letter(max_col)}"
                        f"{translated_row}"
                    )

                    try:
                        ws.merge_cells(
                            original_range
                        )
                    except Exception:
                        pass

                    try:
                        ws.merge_cells(
                            translation_range
                        )
                    except Exception:
                        pass

                else:
                    # Merge nhiều dòng.
                    #
                    # Nếu merge bắt đầu sau row,
                    # insert row làm nó dịch xuống.
                    if min_row > row:

                        new_min_row = (
                            min_row + 1
                        )

                        new_max_row = (
                            max_row + 1
                        )

                    elif (
                        min_row <= row
                        <= max_row
                    ):

                        # Merge cắt qua dòng đang dịch.
                        new_min_row = (
                            min_row
                        )

                        new_max_row = (
                            max_row + 1
                        )

                    else:

                        new_min_row = (
                            min_row
                        )

                        new_max_row = (
                            max_row
                        )

                    new_range = (
                        f"{get_column_letter(min_col)}"
                        f"{new_min_row}:"
                        f"{get_column_letter(max_col)}"
                        f"{new_max_row}"
                    )

                    try:
                        ws.merge_cells(
                            new_range
                        )
                    except Exception:
                        pass

        restore_column_dimensions(
            ws,
            column_dimensions,
        )

    return wb


# ============================================================
# EXCEL SNAPSHOT
# ============================================================

def workbook_info(wb):
    sheets = len(wb.sheetnames)

    total_rows = 0
    total_cells = 0

    for ws in wb.worksheets:

        total_rows += ws.max_row
        total_cells += (
            ws.max_row
            * ws.max_column
        )

    return (
        sheets,
        total_rows,
        total_cells,
    )


# ============================================================
# IMAGE FONT
# ============================================================

def image_font(
    size,
    bold=False,
):
    candidates = []

    if os.name == "nt":

        if bold:
            candidates.extend(
                [
                    r"C:\Windows\Fonts\msyhbd.ttc",
                    r"C:\Windows\Fonts\simhei.ttf",
                    r"C:\Windows\Fonts\arialbd.ttf",
                ]
            )

        candidates.extend(
            [
                r"C:\Windows\Fonts\msyh.ttc",
                r"C:\Windows\Fonts\arial.ttf",
                r"C:\Windows\Fonts\tahoma.ttf",
            ]
        )

    candidates.extend(
        [
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
    )

    for path in candidates:

        if os.path.exists(path):

            try:
                return ImageFont.truetype(
                    path,
                    size=size,
                )
            except Exception:
                pass

    return ImageFont.load_default()


# ============================================================
# IMAGE OCR
# ============================================================

def ocr_image_with_gemini(
    image_bytes,
    mime_type,
):
    """
    Gemini OCR.

    Toạ độ normalized 0-1000.
    """

    prompt = """
Analyze the image as a professional OCR/document-layout system.

Detect ALL readable Chinese or Vietnamese text.

Return JSON only.

Required format:

{
  "items": [
    {
      "id": 1,
      "original_text": "text",
      "x1": 100,
      "y1": 100,
      "x2": 500,
      "y2": 150
    }
  ]
}

Coordinates:
- normalized from 0 to 1000
- x1 = left
- y1 = top
- x2 = right
- y2 = bottom

Rules:

1. Detect every readable text line.
2. Preserve the exact original text.
3. Do not translate.
4. Do not invent text.
5. Include text in tables/forms.
6. Include Chinese and Vietnamese.
7. Ignore purely decorative elements.
8. Keep each logical text line separate.
9. Preserve reading order.
10. Return JSON only.
"""

    raw, used_model = gemini_generate(
        prompt=prompt,
        model=model_choice,
        image=image_bytes,
        mime_type=mime_type,
    )

    data = extract_json(raw)

    items = data.get(
        "items",
        [],
    )

    if not isinstance(
        items,
        list,
    ):
        raise ValueError(
            "Gemini OCR trả dữ liệu không hợp lệ."
        )

    cleaned = []

    for index, item in enumerate(
        items,
        start=1,
    ):

        if not isinstance(
            item,
            dict,
        ):
            continue

        text = clean_text(
            item.get(
                "original_text",
                "",
            )
        )

        if not text:
            continue

        try:

            x1 = float(
                item.get(
                    "x1",
                    0,
                )
            )

            y1 = float(
                item.get(
                    "y1",
                    0,
                )
            )

            x2 = float(
                item.get(
                    "x2",
                    0,
                )
            )

            y2 = float(
                item.get(
                    "y2",
                    0,
                )
            )

        except Exception:

            continue

        cleaned.append(
            {
                "id": index,
                "original_text": text,
                "x1": max(
                    0,
                    min(1000, x1),
                ),
                "y1": max(
                    0,
                    min(1000, y1),
                ),
                "x2": max(
                    0,
                    min(1000, x2),
                ),
                "y2": max(
                    0,
                    min(1000, y2),
                ),
            }
        )

    return (
        cleaned,
        used_model,
    )


# ============================================================
# IMAGE TRANSLATION
# ============================================================

def translate_image_items(
    items,
):
    valid_items = []

    for item in items:

        text = clean_text(
            item.get(
                "original_text",
                "",
            )
        )

        if not text:
            continue

        if not should_translate_cell(
            text
        ):
            continue

        valid_items.append(
            {
                "id": int(
                    item["id"]
                ),
                "text": text,
            }
        )

    if not valid_items:
        return (
            {},
            PRIMARY_MODEL,
        )

    return translate_batch_with_gemini(
        valid_items
    )


# ============================================================
# IMAGE TEXT WRAPPING
# ============================================================

def wrap_text(
    draw,
    text,
    font,
    max_width,
):
    if not text:
        return [""]

    text = str(text)

    # Tiếng Trung thường không có space.
    if " " not in text:

        lines = []
        current = ""

        for char in text:

            candidate = (
                current + char
            )

            bbox = draw.textbbox(
                (0, 0),
                candidate,
                font=font,
            )

            width = (
                bbox[2] - bbox[0]
            )

            if width <= max_width:

                current = candidate

            else:

                if current:
                    lines.append(
                        current
                    )

                current = char

        if current:
            lines.append(
                current
            )

        return lines or [""]

    # Tiếng Việt / text có khoảng trắng
    words = text.split()

    lines = []
    current = ""

    for word in words:

        candidate = (
            word
            if not current
            else current + " " + word
        )

        bbox = draw.textbbox(
            (0, 0),
            candidate,
            font=font,
        )

        width = (
            bbox[2] - bbox[0]
        )

        if width <= max_width:

            current = candidate

        else:

            if current:
                lines.append(
                    current
                )

            current = word

    if current:
        lines.append(
            current
        )

    return lines or [""]


# ============================================================
# IMAGE - DRAW TRANSLATION
# ============================================================

def draw_translation_block(
    draw,
    x,
    y,
    width,
    chinese_text,
    vietnamese_text,
    font_size,
):
    """
    Vẽ:

        中文
        Tiếng Việt

    """

    padding_x = 8
    padding_y = 6

    usable_width = max(
        100,
        width - padding_x * 2,
    )

    chinese_font = image_font(
        max(
            14,
            font_size,
        )
    )

    vietnamese_font = image_font(
        max(
            13,
            int(
                font_size * 0.9
            ),
        )
    )

    chinese_lines = wrap_text(
        draw,
        chinese_text,
        chinese_font,
        usable_width,
    )

    vietnamese_lines = wrap_text(
        draw,
        vietnamese_text,
        vietnamese_font,
        usable_width,
    )

    all_lines = []

    for line in chinese_lines:

        all_lines.append(
            (
                line,
                chinese_font,
            )
        )

    for line in vietnamese_lines:

        all_lines.append(
            (
                line,
                vietnamese_font,
            )
        )

    line_heights = []

    for line, font in all_lines:

        bbox = draw.textbbox(
            (0, 0),
            line,
            font=font,
        )

        line_heights.append(
            max(
                16,
                bbox[3] - bbox[1],
            )
        )

    content_height = sum(
        line_heights
    )

    box_height = (
        content_height
        + padding_y * 2
        + 8
    )

    box_width = min(
        width,
        max(
            width,
            150,
        ),
    )

    # Background
    draw.rounded_rectangle(
        (
            x,
            y,
            x + box_width,
            y + box_height,
        ),
        radius=5,
        fill=(255, 255, 255),
        outline=(0, 0, 0),
        width=1,
    )

    current_y = (
        y + padding_y
    )

    for line_index, (
        line,
        font,
    ) in enumerate(
        all_lines
    ):

        draw.text(
            (
                x + padding_x,
                current_y,
            ),
            line,
            font=font,
            fill=(0, 0, 0),
        )

        current_y += (
            line_heights[
                line_index
            ]
            + 2
        )

    return box_height


# ============================================================
# IMAGE PROCESS
# ============================================================

def process_image(
    image_bytes,
    mime_type,
):
    """
    Nguyên tắc:

    - Không xóa ảnh gốc.
    - Không ghi đè text gốc.
    - Tạo vùng dịch ngay bên dưới text.
    - Nếu phía dưới không đủ chỗ, mở rộng canvas.
    - Trong vùng dịch luôn:
          Trung
          Việt
    """

    image = Image.open(
        io.BytesIO(
            image_bytes
        )
    ).convert("RGB")

    width, height = image.size

    # ========================================================
    # OCR
    # ========================================================

    ocr_items, ocr_model = (
        ocr_image_with_gemini(
            image_bytes,
            mime_type,
        )
    )

    if not ocr_items:
        raise ValueError(
            "Không tìm thấy chữ trong ảnh."
        )

    # ========================================================
    # TRANSLATE
    # ========================================================

    translations, translation_model = (
        translate_image_items(
            ocr_items
        )
    )

    # ========================================================
    # Chuẩn bị dữ liệu
    # ========================================================

    prepared = []

    temp_draw = ImageDraw.Draw(
        image
    )

    max_extra_bottom = 0

    for item in ocr_items:

        item_id = int(
            item["id"]
        )

        original_text = clean_text(
            item["original_text"]
        )

        translated = translations.get(
            item_id,
            "",
        )

        if not translated:
            continue

        x1 = (
            float(item["x1"])
            / 1000
            * width
        )

        y1 = (
            float(item["y1"])
            / 1000
            * height
        )

        x2 = (
            float(item["x2"])
            / 1000
            * width
        )

        y2 = (
            float(item["y2"])
            / 1000
            * height
        )

        # Clamp
        x1 = max(
            0,
            min(
                width - 1,
                x1,
            ),
        )

        y1 = max(
            0,
            min(
                height - 1,
                y1,
            ),
        )

        x2 = max(
            x1 + 20,
            min(
                width,
                x2,
            ),
        )

        y2 = max(
            y1 + 10,
            min(
                height,
                y2,
            ),
        )

        box_width = max(
            120,
            x2 - x1,
        )

        box_height = max(
            12,
            y2 - y1,
        )

        font_size = max(
            14,
            min(
                48,
                int(
                    box_height
                    * 0.75
                ),
            ),
        )

        # ----------------------------------------------------
        # Luôn đặt block dịch ngay dưới text gốc.
        # ----------------------------------------------------

        translation_y = (
            y2 + 8
        )

        # Tạm tính chiều cao.
        test_font = image_font(
            font_size
        )

        chinese_lines = wrap_text(
            temp_draw,
            (
                original_text
                if direction
                == "Trung → Việt"
                else translated
            ),
            test_font,
            max(
                100,
                int(
                    box_width - 16
                ),
            ),
        )

        viet_font = image_font(
            max(
                13,
                int(
                    font_size
                    * 0.9
                ),
            )
        )

        vietnamese_text = (
            translated
            if direction
            == "Trung → Việt"
            else original_text
        )

        chinese_text = (
            original_text
            if direction
            == "Trung → Việt"
            else translated
        )

        vietnamese_lines = wrap_text(
            temp_draw,
            vietnamese_text,
            viet_font,
            max(
                100,
                int(
                    box_width - 16
                ),
            ),
        )

        line_height = max(
            16,
            font_size + 4,
        )

        viet_line_height = max(
            16,
            int(
                font_size
                * 0.9
            ) + 4,
        )

        estimated_height = (
            12
            + len(chinese_lines)
            * line_height
            + 5
            + len(vietnamese_lines)
            * viet_line_height
            + 12
        )

        bottom = (
            translation_y
            + estimated_height
        )

        if bottom > height:

            max_extra_bottom = max(
                max_extra_bottom,
                bottom - height,
            )

        prepared.append(
            {
                "x": int(x1),
                "y": int(y1),
                "x2": int(x2),
                "y2": int(y2),
                "width": int(box_width),
                "font_size": font_size,
                "chinese": chinese_text,
                "vietnamese": vietnamese_text,
                "translation_y": int(
                    translation_y
                ),
            }
        )

    # ========================================================
    # Tạo canvas mới
    # ========================================================

    extra_height = int(
        max(
            100,
            max_extra_bottom + 40,
        )
    )

    output = Image.new(
        "RGB",
        (
            width,
            height + extra_height,
        ),
        (255, 255, 255),
    )

    # Giữ nguyên ảnh gốc.
    output.paste(
        image,
        (0, 0),
    )

    draw = ImageDraw.Draw(
        output
    )

    # ========================================================
    # Vẽ tất cả bản dịch
    # ========================================================

    for item in prepared:

        x = item["x"]
        translation_y = item[
            "translation_y"
        ]

        # Nếu translation_y nằm ngoài
        # canvas thì đưa xuống cuối.
        if (
            translation_y
            >= output.height
        ):

            translation_y = (
                output.height
                - 50
            )

        draw_translation_block(
            draw=draw,
            x=x,
            y=translation_y,
            width=max(
                150,
                item["width"],
            ),
            chinese_text=item[
                "chinese"
            ],
            vietnamese_text=item[
                "vietnamese"
            ],
            font_size=item[
                "font_size"
            ],
        )

    # ========================================================
    # Output PNG
    # ========================================================

    output_buffer = io.BytesIO()

    output.save(
        output_buffer,
        format="PNG",
        optimize=False,
    )

    output_buffer.seek(0)

    return (
        output_buffer.getvalue(),
        ocr_model,
        translation_model,
        len(ocr_items),
    )


# ============================================================
# FILE NAME
# ============================================================

def output_excel_name(
    original_name,
):
    path = Path(
        original_name
    )

    return (
        f"dich_song_ngu_"
        f"{path.stem}"
        f"{path.suffix}"
    )


def output_image_name(
    original_name,
):
    path = Path(
        original_name
    )

    return (
        f"dich_song_ngu_"
        f"{path.stem}.png"
    )


# ============================================================
# UPLOAD
# ============================================================

uploaded_file = st.file_uploader(
    "📁 Tải file lên",
    type=SUPPORTED_EXTENSIONS,
    help=(
        "Hỗ trợ XLSX, XLSM, PNG, JPG, JPEG, WEBP, BMP."
    ),
)


# ============================================================
# API KEY CHECK
# ============================================================

if uploaded_file is not None:

    if not API_KEY:

        st.error(
            "❌ Chưa có Gemini API Key."
        )

        st.info(
            "Bạn có thể nhập API Key ở Sidebar "
            "hoặc cấu hình GEMINI_API_KEY trong Streamlit Secrets."
        )

        st.stop()


# ============================================================
# PROCESS FILE
# ============================================================

if uploaded_file is not None:

    extension = (
        uploaded_file.name
        .rsplit(
            ".",
            1,
        )[-1]
        .lower()
    )

    st.info(
        f"📄 File: **{uploaded_file.name}**"
    )

    # ========================================================
    # EXCEL
    # ========================================================

    if extension in EXCEL_EXTENSIONS:

        st.subheader(
            "📊 Dịch Excel"
        )

        file_bytes = (
            uploaded_file.getvalue()
        )

        # ----------------------------------------------------
        # Preview information
        # ----------------------------------------------------

        try:

            preview_wb = load_workbook(
                io.BytesIO(
                    file_bytes
                ),
                data_only=False,
                keep_vba=(
                    extension
                    == "xlsm"
                ),
            )

            sheets_count = len(
                preview_wb.sheetnames
            )

            st.write(
                f"📑 Số sheet: **{sheets_count}**"
            )

            st.write(
                "Các sheet:"
            )

            for sheet_name in (
                preview_wb.sheetnames
            ):

                st.write(
                    f"• {sheet_name}"
                )

        except Exception as exc:

            st.error(
                "Không thể đọc file Excel."
            )

            st.exception(exc)

            st.stop()

        process_excel_button = (
            st.button(
                "🚀 BẮT ĐẦU DỊCH EXCEL",
                type="primary",
                use_container_width=True,
            )
        )

        if process_excel_button:

            try:

                # ------------------------------------------------
                # Load workbook
                # ------------------------------------------------

                keep_vba = (
                    extension
                    == "xlsm"
                )

                wb = load_workbook(
                    io.BytesIO(
                        file_bytes
                    ),
                    data_only=False,
                    keep_vba=keep_vba,
                )

                original_items = (
                    collect_excel_translation_items(
                        wb
                    )
                )

                total_items = len(
                    original_items
                )

                st.write(
                    f"🔤 Số ô text cần dịch: "
                    f"**{total_items}**"
                )

                if total_items == 0:

                    st.warning(
                        "Không tìm thấy nội dung văn bản cần dịch."
                    )

                    st.stop()

                progress = st.progress(
                    0
                )

                status = st.empty()

                status.info(
                    "🤖 Gemini đang dịch..."
                )

                def update_progress(
                    value
                ):
                    progress.progress(
                        value
                    )

                (
                    translation_map,
                    used_model,
                    translated_count,
                ) = translate_excel_texts(
                    wb,
                    progress_callback=(
                        update_progress
                    ),
                )

                status.success(
                    "✅ Gemini đã dịch xong."
                )

                status.info(
                    "🛠️ Đang tạo Excel song ngữ..."
                )

                # ------------------------------------------------
                # Apply
                # ------------------------------------------------

                wb = apply_excel_translations(
                    wb=wb,
                    translation_map=translation_map,
                    original_items=original_items,
                )

                # ------------------------------------------------
                # Save
                # ------------------------------------------------

                output = io.BytesIO()

                wb.save(output)

                output.seek(0)

                output_bytes = (
                    output.getvalue()
                )

                progress.progress(
                    1.0
                )

                status.success(
                    "🎉 Hoàn tất!"
                )

                # ------------------------------------------------
                # Save session
                # ------------------------------------------------

                st.session_state[
                    "excel_output_bytes"
                ] = output_bytes

                st.session_state[
                    "excel_filename"
                ] = output_excel_name(
                    uploaded_file.name
                )

                st.session_state[
                    "excel_translated_count"
                ] = translated_count

                st.session_state[
                    "excel_used_model"
                ] = used_model

                st.session_state[
                    "excel_sheet_count"
                ] = len(
                    wb.sheetnames
                )

            except Exception as exc:

                st.error(
                    "❌ Xử lý Excel thất bại."
                )

                st.exception(exc)

        # ====================================================
        # DOWNLOAD EXCEL
        # ====================================================

        if (
            "excel_output_bytes"
            in st.session_state
        ):

            st.divider()

            st.subheader(
                "✅ Excel sau khi dịch"
            )

            col1, col2, col3 = st.columns(
                3
            )

            with col1:

                st.metric(
                    "Sheets",
                    st.session_state[
                        "excel_sheet_count"
                    ],
                )

            with col2:

                st.metric(
                    "Ô đã dịch",
                    st.session_state[
                        "excel_translated_count"
                    ],
                )

            with col3:

                st.metric(
                    "Gemini",
                    st.session_state[
                        "excel_used_model"
                    ],
                )

            mime = (
                "application/vnd.ms-excel.sheet.macroEnabled.12"
                if extension
                == "xlsm"
                else
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.download_button(
                label=(
                    "📥 DOWNLOAD EXCEL SONG NGỮ"
                ),
                data=st.session_state[
                    "excel_output_bytes"
                ],
                file_name=st.session_state[
                    "excel_filename"
                ],
                mime=mime,
                type="primary",
                use_container_width=True,
            )

            st.markdown(
                """
                <div class="success-box">
                <b>Quy tắc file xuất:</b><br><br>
                🇨🇳 Tiếng Trung<br>
                🇻🇳 Tiếng Việt<br><br>
                Dòng dịch được đặt ngay bên dưới dòng gốc.
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.warning(
                "Lưu ý: openpyxl có thể không bảo toàn tuyệt đối "
                "các thành phần Excel đặc biệt như PivotTable, "
                "Slicer, SmartArt hoặc một số Drawing phức tạp. "
                "Các thành phần Excel thông thường như font, fill, "
                "border, alignment, merge, row height, column width, "
                "formula, hyperlink và comment được xử lý bảo toàn "
                "ở mức cao nhất có thể."
            )

    # ========================================================
    # IMAGE
    # ========================================================

    elif extension in IMAGE_EXTENSIONS:

        st.subheader(
            "🖼️ Dịch hình ảnh"
        )

        image_bytes = (
            uploaded_file.getvalue()
        )

        mime_map = {
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "webp": "image/webp",
            "bmp": "image/bmp",
        }

        mime_type = mime_map[
            extension
        ]

        # ----------------------------------------------------
        # Preview
        # ----------------------------------------------------

        col1, col2 = st.columns(
            2
        )

        with col1:

            st.markdown(
                "### 🖼️ Ảnh gốc"
            )

            st.image(
                image_bytes,
                use_container_width=True,
            )

        with col2:

            st.markdown(
                "### 📌 Quy tắc"
            )

            st.markdown(
                """
                Ảnh gốc được giữ nguyên.

                Bản dịch được thêm vào vùng bên dưới
                nội dung OCR.

                Kết quả:

                **中文**

                **Tiếng Việt**
                """
            )

        process_image_button = (
            st.button(
                "🚀 OCR + DỊCH HÌNH ẢNH",
                type="primary",
                use_container_width=True,
            )
        )

        if process_image_button:

            try:

                progress = st.progress(
                    0
                )

                status = st.empty()

                status.info(
                    "👁️ Gemini đang nhận dạng chữ..."
                )

                progress.progress(
                    20
                )

                (
                    translated_image,
                    ocr_model,
                    translation_model,
                    ocr_count,
                ) = process_image(
                    image_bytes=image_bytes,
                    mime_type=mime_type,
                )

                progress.progress(
                    100
                )

                status.success(
                    "🎉 OCR + dịch hoàn tất!"
                )

                # ------------------------------------------------
                # Preview
                # ------------------------------------------------

                col1, col2 = st.columns(
                    2
                )

                with col1:

                    st.markdown(
                        "### 🖼️ Ảnh gốc"
                    )

                    st.image(
                        image_bytes,
                        use_container_width=True,
                    )

                with col2:

                    st.markdown(
                        "### 🌐 Ảnh song ngữ"
                    )

                    st.image(
                        translated_image,
                        use_container_width=True,
                    )

                st.success(
                    f"OCR: **{ocr_model}** | "
                    f"Dịch: **{translation_model}** | "
                    f"Vùng chữ: **{ocr_count}**"
                )

                # ------------------------------------------------
                # Session
                # ------------------------------------------------

                st.session_state[
                    "image_output_bytes"
                ] = translated_image

                st.session_state[
                    "image_filename"
                ] = output_image_name(
                    uploaded_file.name
                )

            except Exception as exc:

                st.error(
                    "❌ OCR/Dịch hình ảnh thất bại."
                )

                st.exception(exc)

        # ====================================================
        # DOWNLOAD IMAGE
        # ====================================================

        if (
            "image_output_bytes"
            in st.session_state
        ):

            st.divider()

            st.download_button(
                label=(
                    "📥 DOWNLOAD ẢNH SONG NGỮ"
                ),
                data=st.session_state[
                    "image_output_bytes"
                ],
                file_name=st.session_state[
                    "image_filename"
                ],
                mime="image/png",
                type="primary",
                use_container_width=True,
            )


# ============================================================
# FOOTER
# ============================================================

st.divider()

st.caption(
    "Gemini AI Translator • Excel / Image • Chinese ↔ Vietnamese"
)
